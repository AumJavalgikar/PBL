import openpyxl
from SQL_functions import *
import time

t1 = time.perf_counter()
wbOb = openpyxl.load_workbook('excelsheets/JEE Main Cut offs.xlsx')
sheet = wbOb['Table 1']
all_college_info = []
for row in range(3, sheet.max_row + 1):
    college_info = []
    for column in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(column=column, row=row).value
        try:
            if column > sheet.max_column - 2:
                college_info.append(int(cell_value))
            else:
                college_info.append(cell_value)
        except ValueError:
            print(cell_value)
            cell_value = cell_value[:-1]
            college_info.append(int(cell_value))

    print(f'Finished row {row} out of {sheet.max_row}')
    all_college_info.append(tuple(college_info))
insert_into_JEEtable(all_college_info, 'JEEExcelTable')
print('Successfully inserted all college info into database!')
t2 = time.perf_counter()
print('Finished in', t2-t1, 'seconds')
