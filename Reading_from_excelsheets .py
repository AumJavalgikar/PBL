import openpyxl
from SQL_functions import *

wbOb = openpyxl.load_workbook('pdfs/FE_AI_CAP1-converted.xlsx')
print(wbOb.sheetnames)
print(wbOb['Table 1'])
sheet = wbOb['Table 1']
print(sheet['A2'].value)
all_college_info = []
# for row in range(1, sheet.max_column)
for row in range(3, sheet.max_row + 1):
    college_info = []
    # print('\n')
    for column in range(2, 9):
        if column == 2:
            cutoff = sheet.cell(column=column, row=row).value
            cutoff = cutoff.split(' ')
            merit = cutoff[0]
            score = cutoff[1].strip(')').strip('(')
            college_info.append(merit)
            college_info.append(score)
        else:
            college_info.append(sheet.cell(column=column, row=row).value)
    all_college_info.append(tuple(college_info))
        # print(sheet.cell(column=column, row=row).value, end=' ')
print(sheet.max_row)
print(all_college_info)

insert_into_table(all_college_info, 'CapRound1year18')
