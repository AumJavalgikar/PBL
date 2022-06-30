# This script reads CET cutoff from Excel worksheets and inserts that data into sqlite database.

import openpyxl
from SQL_functions import *
import time


t1 = time.perf_counter()
wbOb = openpyxl.load_workbook('excelsheets/FE_AI_CAP1.xlsx')
sheet = wbOb['Table 1']
all_college_info = []
print('Successfully accessed Table in worksheet \'FE_AI_CAP1.xlsx\'!')
for row in range(3, sheet.max_row + 1):
    college_info = []
    for column in range(2, 9):
        if column == 2:
            cutoff = sheet.cell(column=column, row=row).value
            cutoff = cutoff.split(' ')
            merit = cutoff[0]
            score = cutoff[1].strip(')').strip('(')
            college_info.append(merit)
            college_info.append(score)
        else:
            college_info.append(sheet.cell(column=column, row=row).value)
    all_college_info.append(tuple(college_info))
insert_into_CETtable(all_college_info, 'CETExcelTable')
print('Successfully inserted all college info into database!')
t2 = time.perf_counter()
print('Finished in', t2-t1, 'seconds')
